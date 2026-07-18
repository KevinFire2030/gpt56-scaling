#!/usr/bin/env python3
"""테스트 ①-B '월별 정산서 폼 채우기' — 원장·양식·정답 생성.

산출물:
  test1b/ledger.xlsx        거래 원장 500행 (취소 거래 2건 = 음수 금액 주입)
  test1b/정산서_양식.xlsx    노란 칸 폼 + 헤더 작은 글씨 규칙(십원 미만 절사·취소 제외)
  test1b/_answer/정산서_정답.xlsx + answer.json   봉인 정답 (에이전트 배포 금지)

결정론: random.seed 고정 — 재실행해도 동일 데이터.
"""
import json
import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE = Path(__file__).resolve().parent.parent / "test1b"
ANSWER_DIR = BASE / "_answer"

VENDORS = ["한빛상사", "미래유통", "대성물산", "그린식자재", "서울팩토리",
           "동아컴퍼니", "청록무역", "누리테크"]
ITEMS = ["A등급 원두", "포장 박스", "종이컵 세트", "시럽 원액", "우유 대량",
         "청소용품", "냅킨 묶음", "머신 부품"]
MONTHS = [(2026, 4), (2026, 5), (2026, 6)]

# 취소 거래(음수) 2건 — 숨은 규칙 "취소 거래 제외"의 함정
CANCEL_ROWS = [
    {"date": date(2026, 4, 18), "vendor": "미래유통", "item": "포장 박스 (취소)",
     "qty": -20, "unit_price": 4500},
    {"date": date(2026, 6, 3), "vendor": "대성물산", "item": "시럽 원액 (취소)",
     "qty": -10, "unit_price": 12000},
]


def gen_rows():
    rng = random.Random(20260711)
    rows = []
    for y, m in MONTHS:
        days = (date(y + (m == 12), (m % 12) + 1, 1) - date(y, m, 1)).days
        for _ in range(166):  # 월 ~166행 → 총 ~498 + 취소 2 = 500
            d = date(y, m, rng.randint(1, days))
            vendor = rng.choice(VENDORS)
            item = rng.choice(ITEMS)
            qty = rng.randint(1, 40)
            unit = rng.choice([1200, 2500, 3300, 4500, 5800, 7000, 8800, 12000, 15500])
            rows.append({"date": d, "vendor": vendor, "item": item,
                         "qty": qty, "unit_price": unit})
    rows.extend(CANCEL_ROWS)
    rows.sort(key=lambda r: (r["date"], r["vendor"]))
    return rows


def write_ledger(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "거래원장"
    ws.append(["일자", "거래처", "항목", "수량", "단가", "금액"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        amount = r["qty"] * r["unit_price"]
        ws.append([r["date"].isoformat(), r["vendor"], r["item"],
                   r["qty"], r["unit_price"], amount])
    for col, w in zip("ABCDEF", [12, 14, 18, 8, 10, 12]):
        ws.column_dimensions[col].width = w
    wb.save(BASE / "ledger.xlsx")
    return len(rows)


YELLOW = PatternFill(start_color="FFF2A6", end_color="FFF2A6", fill_type="solid")
THIN = Border(*[Side(style="thin")] * 4)


def build_form(fill_values=None):
    """fill_values=None → 빈 양식 / dict → 정답본."""
    wb = Workbook()
    ws = wb.active
    ws.title = "월별정산서"
    ws["A1"] = "월별 매입 정산서 (2026년 2분기)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    # 숨은 규칙 — 작은 글씨 안내 (①-B의 품질 게이트)
    ws["A2"] = "※ 집계 규칙: 취소 거래(음수 금액)는 제외하고, 각 기입 금액은 십원 미만 절사하여 기입합니다."
    ws["A2"].font = Font(size=8, italic=True, color="808080")
    ws.merge_cells("A2:E2")

    ws.append([])  # row3
    ws.append(["거래처", "4월", "5월", "6월", "분기 합계"])  # row4
    for c in ws[4]:
        c.font = Font(bold=True)
        c.border = THIN
        c.alignment = Alignment(horizontal="center")

    for i, v in enumerate(VENDORS):
        row = 5 + i
        ws.cell(row=row, column=1, value=v).border = THIN
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = YELLOW
            cell.border = THIN
            if fill_values is not None:
                key = (v, ["04", "05", "06", "total"][col - 2])
                cell.value = fill_values[key]
    total_row = 5 + len(VENDORS)
    ws.cell(row=total_row, column=1, value="합계").font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = THIN
    for col in range(2, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = YELLOW
        cell.border = THIN
        if fill_values is not None:
            key = ("합계", ["04", "05", "06", "total"][col - 2])
            cell.value = fill_values[key]
    for col, w in zip("ABCDE", [14, 12, 12, 12, 14]):
        ws.column_dimensions[col].width = w
    return wb


def floor10(v: int) -> int:
    return (v // 10) * 10


def compute_answers(rows):
    """규칙 적용: 취소(음수) 제외 → 셀별 합계 → 십원 미만 절사."""
    raw = {}  # (vendor, month) → sum
    for r in rows:
        amount = r["qty"] * r["unit_price"]
        if amount < 0:
            continue  # 취소 거래 제외 (숨은 규칙)
        key = (r["vendor"], f"{r['date'].month:02d}")
        raw[key] = raw.get(key, 0) + amount
    ans = {}
    for v in VENDORS:
        vt = 0
        for m in ("04", "05", "06"):
            val = floor10(raw.get((v, m), 0))
            ans[(v, m)] = val
            vt += val
        ans[(v, "total")] = floor10(vt)
    for m in ("04", "05", "06", "total"):
        ans[("합계", m)] = floor10(sum(ans[(v, m)] for v in VENDORS))
    return ans


def main():
    BASE.mkdir(parents=True, exist_ok=True)
    ANSWER_DIR.mkdir(exist_ok=True)
    rows = gen_rows()
    n = write_ledger(rows)
    build_form(None).save(BASE / "정산서_양식.xlsx")
    ans = compute_answers(rows)
    build_form(ans).save(ANSWER_DIR / "정산서_정답.xlsx")
    # JSON 정답 (채점 스크립트용)
    (ANSWER_DIR / "answer.json").write_text(json.dumps(
        {f"{k[0]}|{k[1]}": v for k, v in ans.items()},
        ensure_ascii=False, indent=2))
    cancels = [r for r in rows if r["qty"] < 0]
    print(f"원장 {n}행 (취소 {len(cancels)}건: "
          f"{', '.join(r['date'].isoformat() + ' ' + r['vendor'] for r in cancels)})")
    print(f"정답 셀 {len(ans)}개 → {ANSWER_DIR}")


if __name__ == "__main__":
    main()
