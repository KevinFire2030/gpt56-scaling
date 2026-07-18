#!/usr/bin/env python3
"""①-B 정산서 폼 채점기.

사용법: python3 grade_test1b.py --run-dir test1b/<combo>
게이트: G1 구조 불변 · G2 노란 칸 전부 채움 · G3 셀 정확도 ≥95% · G4 숨은 규칙(취소 제외+절사)
evaluator 통과: 구조 OK and 셀 정확도 ≥ 95%
"""
import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from common import emit, fail

BASE = Path(__file__).resolve().parent.parent.parent
ANSWER = json.loads((BASE / "test1b" / "_answer" / "answer.json").read_text())
VENDORS = ["한빛상사", "미래유통", "대성물산", "그린식자재", "서울팩토리",
           "동아컴퍼니", "청록무역", "누리테크"]
MONTH_KEYS = ["04", "05", "06", "total"]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-dir", required=True)
    ap.add_argument("--file", default="정산서_완성.xlsx")
    args = ap.parse_args()
    run = Path(args.run_dir)
    f = run / args.file
    if not f.exists():
        return fail("test1b", f"{args.file} 없음")
    try:
        ws = load_workbook(f, data_only=True)["월별정산서"]
    except Exception as e:
        return fail("test1b", f"열기 실패 또는 시트명 변경: {e}")

    # G1 — 구조 불변 (거래처 행 순서·헤더)
    headers = [ws.cell(row=4, column=c).value for c in range(1, 6)]
    vendors_col = [ws.cell(row=5 + i, column=1).value for i in range(8)]
    structure_ok = (headers == ["거래처", "4월", "5월", "6월", "분기 합계"]
                    and vendors_col == VENDORS
                    and ws.cell(row=13, column=1).value == "합계")

    # G2·G3 — 채움 + 정확도 (36셀)
    labels = VENDORS + ["합계"]
    filled = correct = 0
    wrong = []
    for i, label in enumerate(labels):
        for j, mk in enumerate(MONTH_KEYS):
            v = ws.cell(row=5 + i, column=2 + j).value
            want = ANSWER[f"{label}|{mk}"]
            if v is not None:
                filled += 1
            if v == want:
                correct += 1
            else:
                wrong.append(f"{label}/{mk}: {v} (정답 {want})")
    acc = correct / 36 * 100

    # G4 — 숨은 규칙 진단 (합계행으로 원인 분석)
    diag = "OK"
    if wrong:
        got = ws.cell(row=13, column=5).value  # 합계/total
        want = ANSWER["합계|total"]
        if isinstance(got, (int, float)):
            if got < want:
                diag = "취소 거래 미제외 의심 (합계 부족)"
            elif any(isinstance(ws.cell(row=5+i, column=2+j).value, (int, float))
                     and ws.cell(row=5+i, column=2+j).value % 10 != 0
                     for i in range(9) for j in range(4)):
                diag = "십원 미만 절사 미적용"
            else:
                diag = "집계 오류 (원인 불특정)"

    score = 15 * structure_ok + 15 * (filled / 36) + 55 * (acc / 100) + 15 * (not wrong)
    return emit({
        "test": "test1b", "run_dir": str(run),
        "gates": {"G1_structure": structure_ok, "G2_filled": f"{filled}/36",
                  "G3_cell_acc_pct": round(acc, 2), "G4_rule_diagnosis": diag},
        "wrong_cells": wrong[:8],
        "score_auto": round(max(0, score), 1),
        "evaluator_pass": bool(structure_ok and acc >= 95),
    })


if __name__ == "__main__":
    sys.exit(main())
