from pathlib import Path
import openpyxl

out = Path('정산서_완성.xlsx')
assert out.exists(), '결과 파일이 없습니다.'
assert Path('VERIFICATION.md').exists(), '검증 파일이 없습니다.'
wb = openpyxl.load_workbook(out, data_only=False)
assert wb.sheetnames == ['월별정산서'], wb.sheetnames
ws = wb['월별정산서']
assert ws['A2'].value == '※ 집계 규칙: 취소 거래(음수 금액)는 제외하고, 각 기입 금액은 십원 미만 절사하여 기입합니다.'
yellow = []
for row in ws.iter_rows():
    for cell in row:
        fill = cell.fill
        if fill.fill_type == 'solid' and fill.fgColor.type == 'rgb' and fill.fgColor.rgb == '00FFF2A6':
            yellow.append(cell)
assert len(yellow) == 36, len(yellow)
empty = [c.coordinate for c in yellow if c.value is None or c.value == '']
assert not empty, empty
assert all(isinstance(c.value, (int, float)) for c in yellow), [(c.coordinate, c.value) for c in yellow if not isinstance(c.value, (int, float))]
print('REOPEN_OK')
print('sheet:', wb.sheetnames)
print('yellow_count:', len(yellow))
print('empty_yellow_cells:', len(empty))
print('values:')
for r in range(5, 14):
    print(r, [ws.cell(r, c).value for c in range(1, 6)])
print('files_ok:', out.resolve(), Path('VERIFICATION.md').resolve())
