import openpyxl
for fn in ['ledger.xlsx','정산서_양식.xlsx']:
    wb = openpyxl.load_workbook(fn, data_only=False)
    print('FILE', fn, 'SHEETS', wb.sheetnames)
    for ws in wb.worksheets:
        print('SHEET', ws.title, 'DIM', ws.max_row, ws.max_column)
        for row in ws.iter_rows():
            vals=[]
            for c in row:
                if c.value is not None:
                    vals.append(f'{c.coordinate}={c.value!r}')
            if vals: print('  '+' | '.join(vals))
        if fn=='정산서_양식.xlsx':
            print('FILLS')
            for row in ws.iter_rows():
                for c in row:
                    f=c.fill
                    if f and f.fill_type:
                        print(c.coordinate, repr(c.value), f.fill_type, f.fgColor.type, f.fgColor.rgb, f.fgColor.indexed, f.fgColor.theme, f.fgColor.tint)
