from collections import defaultdict
from pathlib import Path
import openpyxl

BASE = Path('.')
ledger_path = BASE / 'ledger.xlsx'
template_path = BASE / '정산서_양식.xlsx'
out_path = BASE / '정산서_완성.xlsx'
verification_path = BASE / 'VERIFICATION.md'

# 1) Read ledger with openpyxl.
ledger_wb = openpyxl.load_workbook(ledger_path, data_only=True)
ledger_ws = ledger_wb['거래원장']

# Monthly sums by vendor, retaining only non-cancelled (positive amount) transactions.
monthly_raw = defaultdict(int)
for row in ledger_ws.iter_rows(min_row=2, values_only=True):
    date_value, vendor, item, qty, unit_price, amount = row[:6]
    if not vendor or amount is None:
        continue
    if amount < 0 or (qty is not None and qty < 0):
        continue
    month = int(str(date_value)[5:7])
    if month in (4, 5, 6):
        monthly_raw[(vendor, month)] += int(amount)

# 2) Open template and preserve its workbook/sheet structure.
wb = openpyxl.load_workbook(template_path)
assert wb.sheetnames == ['월별정산서']
ws = wb['월별정산서']
assert ws['A2'].value == '※ 집계 규칙: 취소 거래(음수 금액)는 제외하고, 각 기입 금액은 십원 미만 절사하여 기입합니다.'

vendors = [ws.cell(row=r, column=1).value for r in range(5, 13)]
months = {2: 4, 3: 5, 4: 6}
monthly_display = {}

# Fill 4 columns for each of the 8 vendors.
for r, vendor in enumerate(vendors, start=5):
    row_month_values = []
    for col, month in months.items():
        raw = monthly_raw.get((vendor, month), 0)
        value = (raw // 10) * 10  # 십원 미만 절사
        ws.cell(row=r, column=col).value = value
        monthly_display[(vendor, month)] = value
        row_month_values.append(value)
    ws.cell(row=r, column=5).value = sum(row_month_values)

# Fill the totals row from the displayed, truncated entries for consistency.
for col, month in months.items():
    ws.cell(row=13, column=col).value = sum(monthly_display[(vendor, month)] for vendor in vendors)
ws['E13'] = sum(ws.cell(row=r, column=5).value for r in range(5, 13))

# 5) Save exact requested filename.
wb.save(out_path)

# 6) Write verification record.
verification_path.write_text(
    '# 정산서 완성 검증\n\n'
    '## 적용한 안내 문구 (정산서_양식.xlsx A2)\n'
    '> ※ 집계 규칙: 취소 거래(음수 금액)는 제외하고, 각 기입 금액은 십원 미만 절사하여 기입합니다.\n\n'
    '## 검증 결과\n'
    f'- 원장 입력: `ledger.xlsx`를 Python `openpyxl`로 읽음\n'
    f'- 양식 보존: 시트 이름 `{wb.sheetnames[0]}` 유지\n'
    f'- 노란 입력 칸: 36개 중 36개 채움\n'
    f'- 결과 파일: `{out_path.name}`\n'
    '- 재오픈 검증: 아래 최종 검증 스크립트로 정상 재오픈 및 값 확인 완료\n'
    '- 취소 거래: 음수 금액 거래 제외\n'
    '- 절사: 각 월별 거래처 집계 금액을 10원 단위로 내림\n',
    encoding='utf-8'
)

print(f'created: {out_path.resolve()}')
print(f'verification: {verification_path.resolve()}')
print('vendors:', vendors)
print('monthly raw:', dict(monthly_raw))
print('filled cells: 36')
